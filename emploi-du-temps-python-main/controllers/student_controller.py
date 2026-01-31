"""
Contrôleur pour les fonctionnalités étudiants
"""

from datetime import datetime
from database import getConnection

# Jours de la semaine
DAYS = {1: "Lundi", 2: "Mardi", 3: "Mercredi", 4: "Jeudi", 5: "Vendredi"}

class StudentController:
    def __init__(self, user_id):
        """
        Initialise le contrôleur avec l'ID de l'utilisateur étudiant
        """
        self.user_id = user_id
        self.group_id = self._get_student_group()
    
    def _get_student_group(self):
        """
        Trouve le groupe de l'étudiant via la table student_groups.
        
        Returns:
            int or None: ID du groupe ou None si non trouvé
        """
        conn = getConnection()
        cursor = conn.cursor()
        
        # Récupérer le groupe de l'étudiant via la table de liaison
        cursor.execute("""
            SELECT g.id 
            FROM groups g
            JOIN student_groups sg ON g.id = sg.group_id
            WHERE sg.user_id = ? AND g.active = 1
            LIMIT 1
        """, (self.user_id,))
        
        result = cursor.fetchone()
        
        # Fallback: si pas dans student_groups, prendre le premier groupe actif
        if not result:
            cursor.execute("SELECT id FROM groups WHERE active = 1 LIMIT 1")
            result = cursor.fetchone()
        
        conn.close()
        return result['id'] if result else None
    
    def get_group_timetable(self):
        """
        CONSULTER L'EMPLOI DU TEMPS DE SON GROUPE
        Retourne l'emploi du temps complet du groupe
        """
        if not self.group_id:
            return {"success": False, "error": "Groupe non trouvé pour cet étudiant"}
        
        conn = getConnection()
        cursor = conn.cursor()
        
        # Récupérer le nom du groupe
        cursor.execute("SELECT name FROM groups WHERE id = ?", (self.group_id,))
        group = cursor.fetchone()
        group_name = group['name'] if group else "Inconnu"
        
        # Récupérer l'emploi du temps
        query = """
        SELECT 
            t.day,
            t.start_hour,
            t.duration,
            s.name AS subject_name,
            s.code AS subject_code,
            s.type AS subject_type,
            i.name AS instructor_name,
            r.name AS room_name,
            r.type AS room_type
        FROM timetable t
        JOIN subjects s ON t.course_id = s.id
        JOIN instructors i ON t.instructor_id = i.id
        JOIN rooms r ON t.room_id = r.id
        WHERE t.group_id = ?
        ORDER BY t.day, t.start_hour
        """
        
        cursor.execute(query, (self.group_id,))
        timetable_slots = cursor.fetchall()
        conn.close()
        
        # Organiser par jour
        organized = {}
        for day_num, day_name in DAYS.items():
            organized[day_name] = []
        
        for slot in timetable_slots:
            day_name = DAYS.get(slot['day'], 'Inconnu')
            end_hour = slot['start_hour'] + slot['duration']
            slot_info = {
                'jour': day_name,
                'horaire': f"{slot['start_hour']:02d}h-{end_hour:02d}h",
                'matiere': slot['subject_name'],
                'code': slot['subject_code'],
                'type': slot['subject_type'],
                'enseignant': slot['instructor_name'],
                'salle': slot['room_name'],
                'type_salle': slot['room_type']
            }
            organized[day_name].append(slot_info)
        
        return {
            "success": True, 
            "groupe": group_name,
            "emploi_du_temps": organized
        }
    
    def search_free_room(self, day=None, start_hour=None, duration=2):
        """
        RECHERCHER UNE SALLE LIBRE
        Pour travaux de groupe, révisions, etc.
        
        Si jour et heure spécifiés: cherche les salles libres à ce créneau
        Si seulement jour spécifié: montre toutes les salles avec leurs disponibilités
        Si rien spécifié: liste toutes les salles
        """
        conn = getConnection()
        cursor = conn.cursor()
        
        if day and start_hour:
            # Recherche précise pour un créneau
            end_hour = start_hour + duration
            query = """
            SELECT r.name, r.type, r.capacity FROM rooms r
            WHERE r.active = 1 
            AND r.id NOT IN (
                SELECT room_id FROM timetable 
                WHERE day = ? 
                AND (start_hour < ?) AND (? < start_hour + duration)
                UNION
                SELECT room_id FROM reservations 
                WHERE day = ? AND status = 'APPROVED'
                AND (start_hour < ?) AND (? < start_hour + duration)
            )
            ORDER BY r.name
            """
            
            cursor.execute(query, (day, end_hour, start_hour, day, end_hour, start_hour))
            rooms = cursor.fetchall()
            conn.close()
            
            # MODIFICATION: Retourner des noms au lieu d'IDs
            rooms_list = []
            for room in rooms:
                rooms_list.append({
                    'nom': room['name'],
                    'type': room['type'],
                    'capacité': room['capacity'],
                    'horaire': f"{start_hour}h-{end_hour}h"
                })
            
            return {"success": True, "rooms": rooms_list}
        
        elif day:
            # Voir les disponibilités sur toute la journée
            cursor.execute("SELECT * FROM rooms WHERE active = 1 ORDER BY name")
            all_rooms = cursor.fetchall()
            
            rooms_with_schedule = []
            for room in all_rooms:
                # Récupérer les créneaux occupés
                cursor.execute("""
                    SELECT start_hour, duration 
                    FROM timetable 
                    WHERE room_id = ? AND day = ?
                    UNION
                    SELECT start_hour, duration 
                    FROM reservations 
                    WHERE room_id = ? AND day = ? AND status = 'APPROVED'
                    ORDER BY start_hour
                """, (room['id'], day, room['id'], day))
                
                occupied = cursor.fetchall()
                
                # Calculer les créneaux libres (8h-18h)
                free_slots = []
                current = 8  # Début de journée
                
                for slot in occupied:
                    slot_start = slot['start_hour']
                    slot_end = slot_start + slot['duration']
                    
                    if current < slot_start:
                        free_slots.append(f"{current}h-{slot_start}h")
                    
                    current = max(current, slot_end)
                
                if current < 18:
                    free_slots.append(f"{current}h-18h")
                
                rooms_with_schedule.append({
                    'nom': room['name'],
                    'type': room['type'],
                    'capacité': room['capacity'],
                    'creneaux_libres': free_slots
                })
            
            conn.close()
            return {"success": True, "rooms": rooms_with_schedule}
        
        else:
            # Lister toutes les salles (retourner des noms)
            cursor.execute("SELECT name, type, capacity, equipments FROM rooms WHERE active = 1 ORDER BY name")
            rooms = cursor.fetchall()
            conn.close()
            
            rooms_list = []
            for room in rooms:
                rooms_list.append({
                    'nom': room['name'],
                    'type': room['type'],
                    'capacité': room['capacity'],
                    'équipements': room['equipments'] or "Aucun"
                })
            
            return {"success": True, "rooms": rooms_list}
    
    def get_today_schedule(self):
        """
        VOIR L'EMPLOI DU TEMPS D'AUJOURD'HUI
        Version simplifiée pour consultation rapide
        """
        if not self.group_id:
            return {"success": False, "error": "Groupe non trouvé"}
        
        # Jour actuel (1=Lundi, 5=Vendredi)
        today = datetime.now().weekday() + 1
        
        conn = getConnection()
        cursor = conn.cursor()
        
        query = """
        SELECT 
            t.start_hour,
            t.duration,
            s.name AS subject_name,
            i.name AS instructor_name,
            r.name AS room_name
        FROM timetable t
        JOIN subjects s ON t.course_id = s.id
        JOIN instructors i ON t.instructor_id = i.id
        JOIN rooms r ON t.room_id = r.id
        WHERE t.group_id = ? AND t.day = ?
        ORDER BY t.start_hour
        """
        
        cursor.execute(query, (self.group_id, today))
        today_schedule = cursor.fetchall()
        conn.close()
        
        schedule_list = []
        for slot in today_schedule:
            end_hour = slot['start_hour'] + slot['duration']
            schedule_list.append({
                'horaire': f"{slot['start_hour']:02d}h-{end_hour:02d}h",
                'matiere': slot['subject_name'],
                'enseignant': slot['instructor_name'],
                'salle': slot['room_name']
            })
        
        return {
            "success": True,
            "jour": DAYS.get(today, "Aujourd'hui"),
            "cours": schedule_list,
            "nombre_cours": len(schedule_list)
        }

    # -------------------------------------------------------------------------
    # EXPORTS
    # -------------------------------------------------------------------------
    
    def export_my_timetable_pdf(self, filename="Mon_Emploi_du_Temps.pdf"):
        """Export student timetable to PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            import os
        except ImportError:
            return "Erreur: reportlab non installé."

        # Setup paths
        base_dir = os.getcwd()
        exports_dir = os.path.join(base_dir, "exports")
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
        unique_filename = os.path.join(exports_dir, filename)

        doc = SimpleDocTemplate(unique_filename, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        # Styles
        style_header = ParagraphStyle('Header', fontSize=12, leading=14, alignment=1)
        style_title = ParagraphStyle('Title', fontSize=16, leading=20, alignment=1, spaceAfter=20, fontName='Helvetica-Bold')

        # Header
        elements.append(Paragraph("<b>UNIVERSITÉ ABDELMALEK ESSAADI - FST TANGER</b>", style_header))
        elements.append(Spacer(1, 20))
        
        # Get Group Name
        conn = getConnection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM groups WHERE id = ?", (self.group_id,))
        res = cursor.fetchone()
        group_name = res['name'] if res else "Groupe Inconnu"
        
        elements.append(Paragraph(f"Mon Emploi du Temps - Étudiant ({group_name})", style_title))

        # Data Matrix
        time_slots = ["08h00-09h30", "09h00-10h30", "10h45-12h15", "12h30-14h00", "14h15-15h45", "16h00-17h30"]
        data = [["JOURS"] + time_slots]
        days_list = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]
        DAYS_MAPPING = {"LUNDI": 1, "MARDI": 2, "MERCREDI": 3, "JEUDI": 4, "VENDREDI": 5, "SAMEDI": 6}
        SLOT_TO_HOUR = {"08h00-09h30": 8, "09h00-10h30": 9, "10h45-12h15": 10, "12h30-14h00": 12, "14h15-15h45": 14, "16h00-17h30": 16}

        for day_name in days_list:
            row = [day_name]
            day_idx = DAYS_MAPPING.get(day_name)
            
            for slot in time_slots:
                if day_name == "VENDREDI" and slot == "14h15-15h45":
                    start_h = 15
                else:
                    start_h = SLOT_TO_HOUR.get(slot, int(slot.split('h')[0]))
                
                cursor.execute("""
                    SELECT s.name as subject, r.name as room, i.name as instructor
                    FROM timetable t
                    JOIN subjects s ON t.course_id = s.id
                    JOIN rooms r ON t.room_id = r.id
                    JOIN instructors i ON t.instructor_id = i.id
                    WHERE t.group_id = ? AND t.day = ? AND t.start_hour = ?
                """, (self.group_id, day_idx, start_h))
                
                results = cursor.fetchall()
                if results:
                    cell_text = "\n".join([f"{r['subject']}\n{r['room']}\n({r['instructor']})" for r in results])
                    row.append(cell_text)
                else:
                    row.append("")
            data.append(row)
        
        conn.close()

        # Table Style
        table = Table(data, colWidths=[80] + [110] * len(time_slots))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (0, -1), colors.whitesmoke),
        ]))
        elements.append(table)
        
        doc.build(elements)
        return f"PDF exporté vers: {unique_filename}"

    def export_my_timetable_excel(self, filename="Mon_Emploi_du_Temps.xlsx"):
        """Export student timetable to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            import os
        except ImportError:
            return "Erreur: openpyxl non installé"
            
        # Setup paths
        base_dir = os.getcwd()
        exports_dir = os.path.join(base_dir, "exports")
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
        unique_filename = os.path.join(exports_dir, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mon Emploi du Temps"
        
        # Styles
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_header = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Headers
        time_slots = ["08h00-09h30", "09h00-10h30", "10h45-12h15", "12h30-14h00", "14h15-15h45", "16h00-17h30"]
        ws.append(["JOURS"] + time_slots)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill_header
            cell.alignment = center_align
            cell.border = thin_border
            
        # Data
        days_list = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]
        DAYS_MAPPING = {"LUNDI": 1, "MARDI": 2, "MERCREDI": 3, "JEUDI": 4, "VENDREDI": 5, "SAMEDI": 6}
        SLOT_TO_HOUR = {"08h00-09h30": 8, "09h00-10h30": 9, "10h45-12h15": 10, "12h30-14h00": 12, "14h15-15h45": 14, "16h00-17h30": 16}
        
        conn = getConnection()
        cursor = conn.cursor()
        
        for day_name in days_list:
            row_data = [day_name]
            day_idx = DAYS_MAPPING.get(day_name)
            
            for slot in time_slots:
                if day_name == "VENDREDI" and slot == "14h15-15h45":
                    start_h = 15
                else:
                    start_h = SLOT_TO_HOUR.get(slot, int(slot.split('h')[0]))
                
                cursor.execute("""
                    SELECT s.name as subject, r.name as room, i.name as instructor
                    FROM timetable t
                    JOIN subjects s ON t.course_id = s.id
                    JOIN rooms r ON t.room_id = r.id
                    JOIN instructors i ON t.instructor_id = i.id
                    WHERE t.group_id = ? AND t.day = ? AND t.start_hour = ?
                """, (self.group_id, day_idx, start_h))
                
                results = cursor.fetchall()
                if results:
                    text = "\n".join([f"{r['subject']}\n{r['room']}\n{r['instructor']}" for r in results])
                    row_data.append(text)
                else:
                    row_data.append("")
            
            ws.append(row_data)
        
        conn.close()
        
        # Formatting rows
        for row in ws.iter_rows(min_row=2, max_row=len(days_list)+1):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border
        
        # Dimensions
        ws.column_dimensions['A'].width = 15
        for col in range(2, 7):
             ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
             
        wb.save(unique_filename)
        return f"Excel exporté vers: {unique_filename}"

    def export_my_timetable_image(self, filename="Mon_Emploi_du_Temps.png"):
        """Export student timetable to Image (PNG)"""
        try:
            from PIL import Image, ImageDraw, ImageFont
            import os
        except ImportError:
            return "Erreur: Pillow non installé"
            
        # Setup paths
        base_dir = os.getcwd()
        exports_dir = os.path.join(base_dir, "exports")
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
        unique_filename = os.path.join(exports_dir, filename)
        
        # Config
        time_slots = ["08h00-09h30", "09h00-10h30", "10h45-12h15", "12h30-14h00", "14h15-15h45", "16h00-17h30"]
        days_list = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]
        DAYS_MAPPING = {"LUNDI": 1, "MARDI": 2, "MERCREDI": 3, "JEUDI": 4, "VENDREDI": 5, "SAMEDI": 6}
        SLOT_TO_HOUR = {"08h00-09h30": 8, "09h00-10h30": 9, "10h45-12h15": 10, "12h30-14h00": 12, "14h15-15h45": 14, "16h00-17h30": 16}
        
        cell_width = 140 # Slightly reduced
        cell_height = 90
        header_height = 80
        col_header_height = 40
        day_col_width = 100
        
        img_width = day_col_width + (len(time_slots) * cell_width) + 20
        img_height = header_height + col_header_height + (len(days_list) * cell_height) + 20
        
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Fonts
        try:
            font_title = ImageFont.truetype("arial.ttf", 20)
            font_header = ImageFont.truetype("arial.ttf", 12)
            font_cell = ImageFont.truetype("arial.ttf", 10)
        except:
            font_title = ImageFont.load_default()
            font_header = ImageFont.load_default()
            font_cell = ImageFont.load_default()
            
        # Title
        draw.text((20, 20), "Mon Emploi du Temps - FST Tanger", fill='black', font=font_title)
        
        start_x = 10
        start_y = header_height
        
        # Header Row
        draw.rectangle([start_x, start_y, start_x + day_col_width, start_y + col_header_height], fill='#4472C4', outline='black')
        draw.text((start_x + 10, start_y + 10), "JOURS", fill='white', font=font_header)
        
        for i, slot in enumerate(time_slots):
            x = start_x + day_col_width + (i * cell_width)
            draw.rectangle([x, start_y, x + cell_width, start_y + col_header_height], fill='#4472C4', outline='black')
            draw.text((x + 10, start_y + 10), slot, fill='white', font=font_header)
            
        # Data Rows
        conn = getConnection()
        cursor = conn.cursor()
        
        for row_idx, day_name in enumerate(days_list):
            y = start_y + col_header_height + (row_idx * cell_height)
            
            # Day Column
            draw.rectangle([start_x, y, start_x + day_col_width, y + cell_height], fill='#D9E2F3', outline='black')
            draw.text((start_x + 10, y + 35), day_name, fill='black', font=font_header)
            
            day_idx = DAYS_MAPPING.get(day_name)
            
            for col_idx, slot in enumerate(time_slots):
                x = start_x + day_col_width + (col_idx * cell_width)
                
                if day_name == "VENDREDI" and slot == "14h15-15h45":
                    start_h = 15
                else:
                    start_h = SLOT_TO_HOUR.get(slot, int(slot.split('h')[0]))
                    
                cursor.execute("""
                    SELECT s.name as subject, r.name as room, i.name as instructor
                    FROM timetable t
                    JOIN subjects s ON t.course_id = s.id
                    JOIN rooms r ON t.room_id = r.id
                    JOIN instructors i ON t.instructor_id = i.id
                    WHERE t.group_id = ? AND t.day = ? AND t.start_hour = ?
                """, (self.group_id, day_idx, start_h))
                
                results = cursor.fetchall()
                draw.rectangle([x, y, x + cell_width, y + cell_height], outline='black')
                
                if results:
                    text_lines = []
                    for r in results:
                        text_lines.append(f"{r['subject']}")
                        text_lines.append(f"{r['room']} - {r['instructor']}")
                    
                    text_y = y + 10
                    for line in text_lines:
                        draw.text((x + 5, text_y), line, fill='black', font=font_cell)
                        text_y += 15
        
        conn.close()
        img.save(unique_filename)
        return f"Image exportée vers: {unique_filename}"