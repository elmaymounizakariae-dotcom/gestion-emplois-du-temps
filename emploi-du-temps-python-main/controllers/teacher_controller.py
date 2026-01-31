"""
Contrôleur pour les fonctionnalités enseignants
"""

import sqlite3
from datetime import datetime
from database import getConnection

# Jours de la semaine (copié de database.py pour éviter l'import circulaire)
DAYS = {1: "Lundi", 2: "Mardi", 3: "Mercredi", 4: "Jeudi", 5: "Vendredi"}

class TeacherController:
    def __init__(self, user_id):
        """
        Initialise le contrôleur avec l'ID de l'utilisateur enseignant
        """
        self.user_id = user_id
        self.instructor_id = self._get_instructor_id()
    
    def _get_instructor_id(self):
        """Récupère l'ID de l'instructeur"""
        conn = getConnection()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM instructors WHERE user_id = ?", (self.user_id,))
        result = cursor.fetchone()
        conn.close()
        return result['id'] if result else None
    
    def get_teacher_timetable(self):
        """
        CONSULTER SON EMPLOI DU TEMPS
        Retourne l'emploi du temps personnalisé de l'enseignant
        """
        if not self.instructor_id:
            return {"success": False, "error": "Enseignant non trouvé"}
        
        conn = getConnection()
        cursor = conn.cursor()
        
        query = """
        SELECT 
            t.day,
            t.start_hour,
            t.duration,
            s.name AS subject_name,
            s.code AS subject_code,
            g.name AS group_name,
            r.name AS room_name,
            r.type AS room_type
        FROM timetable t
        JOIN subjects s ON t.course_id = s.id
        JOIN groups g ON t.group_id = g.id
        JOIN rooms r ON t.room_id = r.id
        WHERE t.instructor_id = ?
        ORDER BY t.day, t.start_hour
        """
        
        cursor.execute(query, (self.instructor_id,))
        timetable_slots = cursor.fetchall()
        conn.close()
        
        # Organiser par jour
        organized_timetable = {}
        for day_num, day_name in DAYS.items():
            organized_timetable[day_name] = []
        
        for slot in timetable_slots:
            day_name = DAYS.get(slot['day'], 'Inconnu')
            end_hour = slot['start_hour'] + slot['duration']
            slot_info = {
                'jour': day_name,
                'heure': f"{slot['start_hour']:02d}h-{end_hour:02d}h",
                'matiere': slot['subject_name'],
                'code': slot['subject_code'],
                'groupe': slot['group_name'],
                'salle': slot['room_name'],
                'type_salle': slot['room_type']
            }
            organized_timetable[day_name].append(slot_info)
        
        return {"success": True, "timetable": organized_timetable}
    
    # MODIFICATION: Cette méthode prend maintenant des noms au lieu d'IDs
    def submit_reservation(self, room_name, group_name, day, start_hour, duration, reason=""):
        """
        SOUMETTRE UNE RÉSERVATION PONCTUELLE
        Demande de réservation de salle pour séance de rattrapage, réunion, etc.
        """
        # Validation basique
        if day < 1 or day > 5:
            return {"success": False, "message": "Jour invalide (1-5)"}
        if start_hour < 8 or start_hour > 18:
            return {"success": False, "message": "Heure invalide (8-18h)"}
        if duration < 1 or duration > 4:
            return {"success": False, "message": "Durée invalide (1-4h)"}
        
        if not self.instructor_id:
            return {"success": False, "message": "Enseignant non trouvé"}
        
        conn = getConnection()
        cursor = conn.cursor()
        
        # Récupérer l'ID de la salle par son nom
        cursor.execute("SELECT id FROM rooms WHERE name = ? AND active = 1", (room_name,))
        room = cursor.fetchone()
        if not room:
            conn.close()
            return {"success": False, "message": f"Salle '{room_name}' non trouvée"}
        room_id = room['id']
        
        # Récupérer l'ID du groupe par son nom
        cursor.execute("SELECT id FROM groups WHERE name = ? AND active = 1", (group_name,))
        group = cursor.fetchone()
        if not group:
            conn.close()
            return {"success": False, "message": f"Groupe '{group_name}' non trouvé"}
        group_id = group['id']
        
        # Vérifier si la salle est disponible
        conflict = self._check_room_availability(room_id, day, start_hour, duration)
        if conflict:
            conn.close()
            return {"success": False, "message": conflict}
        
        try:
            cursor.execute("""
                INSERT INTO reservations 
                (instructor_id, room_id, group_id, day, start_hour, duration, reason, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'PENDING')
            """, (self.instructor_id, room_id, group_id, day, start_hour, duration, reason))
            
            conn.commit()
            reservation_id = cursor.lastrowid
            conn.close()
            
            return {
                "success": True, 
                "message": f"Demande de réservation soumise avec succès (ID: {reservation_id})",
                "reservation_id": reservation_id
            }
        except sqlite3.IntegrityError as e:
            conn.close()
            return {"success": False, "message": f"Erreur: {str(e)}"}
    
    def _check_room_availability(self, room_id, day, start_hour, duration):
        """Vérifie si une salle est disponible à un créneau donné"""
        conn = getConnection()
        cursor = conn.cursor()
        end_hour = start_hour + duration
        
        # Vérifier dans l'emploi du temps
        cursor.execute("""
            SELECT id FROM timetable 
            WHERE room_id = ? AND day = ? 
            AND (start_hour < ?) AND (? < start_hour + duration)
        """, (room_id, day, end_hour, start_hour))
        
        if cursor.fetchone():
            conn.close()
            return "Salle déjà occupée dans l'emploi du temps"
        
        # Vérifier dans les réservations approuvées
        cursor.execute("""
            SELECT id FROM reservations 
            WHERE room_id = ? AND day = ? AND status = 'APPROVED'
            AND (start_hour < ?) AND (? < start_hour + duration)
        """, (room_id, day, end_hour, start_hour))
        
        if cursor.fetchone():
            conn.close()
            return "Salle déjà réservée à ce créneau"
        
        conn.close()
        return None
    
    def declare_unavailability(self, day, start_hour, duration, reason=""):
        """
        DÉCLARER UNE INDISPONIBILITÉ
        L'enseignant signale qu'il n'est pas disponible à un créneau
        """
        # Validation
        if day < 1 or day > 5:
            return {"success": False, "message": "Jour invalide (1-5)"}
        if start_hour < 8 or start_hour > 18:
            return {"success": False, "message": "Heure invalide (8-18h)"}
        
        if not self.instructor_id:
            return {"success": False, "message": "Enseignant non trouvé"}
        
        conn = getConnection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                INSERT INTO teacher_unavailability 
                (instructor_id, day, start_hour, duration, reason)
                VALUES (?, ?, ?, ?, ?)
            """, (self.instructor_id, day, start_hour, duration, reason))
            
            conn.commit()
            conn.close()
            
            # Mettre à jour les indisponibilités dans la table instructors
            self._update_unavailable_slots()
            
            return {"success": True, "message": "Indisponibilité déclarée avec succès"}
        except Exception as e:
            conn.close()
            return {"success": False, "message": f"Erreur: {str(e)}"}
    
    def _update_unavailable_slots(self):
        """Met à jour le champ unavailable_slots dans instructors"""
        conn = getConnection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT day, start_hour, duration 
            FROM teacher_unavailability 
            WHERE instructor_id = ?
        """, (self.instructor_id,))
        
        slots = cursor.fetchall()
        
        # Formater les créneaux
        formatted = []
        for slot in slots:
            day_name = DAYS.get(slot['day'], f"Jour{slot['day']}")
            start = slot['start_hour']
            end = start + slot['duration']
            formatted.append(f"{day_name}_{start:02d}-{end:02d}")
        
        # Mettre à jour
        cursor.execute("""
            UPDATE instructors 
            SET unavailable_slots = ?
            WHERE id = ?
        """, (','.join(formatted), self.instructor_id))
        
        conn.commit()
        conn.close()
    
    def search_available_room(self, day, start_hour, duration=2, min_capacity=30):
        """
        RECHERCHER UNE SALLE VACANTE
        Recherche selon critères (horaire, capacité, équipement)
        """
        # Validation
        if day < 1 or day > 5:
            return {"success": False, "message": "Jour invalide", "rooms": []}
        
        conn = getConnection()
        cursor = conn.cursor()
        end_hour = start_hour + duration
        
        query = """
        SELECT r.* FROM rooms r
        WHERE r.active = 1 
        AND r.capacity >= ?
        AND r.id NOT IN (
            SELECT room_id FROM timetable 
            WHERE day = ? 
            AND (start_hour < ?) AND (? < start_hour + duration)
            UNION
            SELECT room_id FROM reservations 
            WHERE day = ? AND status = 'APPROVED'
            AND (start_hour < ?) AND (? < start_hour + duration)
        )
        ORDER BY r.name
        """
        
        params = [min_capacity, day, end_hour, start_hour, day, end_hour, start_hour]
        cursor.execute(query, params)
        rooms = cursor.fetchall()
        conn.close()
        
        # Formater les résultats
        rooms_list = []
        for room in rooms:
            rooms_list.append({
                'nom': room['name'],
                'type': room['type'],
                'capacité': room['capacity'],
                'équipements': room['equipments'],
                'disponible': True
            })
        
        return {"success": True, "rooms": rooms_list, "count": len(rooms_list)}
    
    def get_reservation_status(self):
        """Voir le statut des réservations soumises"""
        if not self.instructor_id:
            return {"success": False, "reservations": []}
        
        conn = getConnection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                r.id, r.day, r.start_hour, r.duration,
                ro.name AS salle, r.reason, r.status,
                r.created_at AS date_soumission
            FROM reservations r
            LEFT JOIN rooms ro ON r.room_id = ro.id
            WHERE r.instructor_id = ?
            ORDER BY r.created_at DESC
        """, (self.instructor_id,))
        
        reservations = cursor.fetchall()
        conn.close()
        
        # Formater
        formatted = []
        for res in reservations:
            day_name = DAYS.get(res['day'], f"Jour{res['day']}")
            end_hour = res['start_hour'] + res['duration']
            formatted.append({
                'id': res['id'],
                'jour': day_name,
                'horaire': f"{res['start_hour']}h-{end_hour}h",
                'salle': res['salle'],
                'motif': res['reason'],
                'statut': res['status'],
                'soumis_le': res['date_soumission']
            })
        
        return {"success": True, "reservations": formatted}

    # -------------------------------------------------------------------------
    # EXPORTS
    # -------------------------------------------------------------------------
    
    def export_my_timetable_pdf(self, filename="Mon_Planning_Enseignant.pdf"):
        """Export teacher timetable to PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            import os
        except ImportError:
            return "Erreur: reportlab non installé."

        # Setup paths
        base_dir = os.getcwd()
        exports_dir = os.path.join(base_dir, "exports")
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
        unique_filename = os.path.join(exports_dir, filename)

        doc = SimpleDocTemplate(unique_filename, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        # Styles
        style_header = ParagraphStyle('Header', fontSize=12, leading=14, alignment=1)
        style_title = ParagraphStyle('Title', fontSize=16, leading=20, alignment=1, spaceAfter=20, fontName='Helvetica-Bold')

        # Header
        elements.append(Paragraph("<b>UNIVERSITÉ ABDELMALEK ESSAADI - FST TANGER</b>", style_header))
        elements.append(Spacer(1, 20))
        
        # Get Instructor Name
        conn = getConnection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM instructors WHERE id = ?", (self.instructor_id,))
        res = cursor.fetchone()
        name = res['name'] if res else "Enseignant"
        
        elements.append(Paragraph(f"Mon Emploi du Temps - Enseignant: {name}", style_title))

        # Data Matrix
        time_slots = ["09h00-10h30", "10h45-12h15", "12h30-14h00", "14h15-15h45", "16h00-17h30"]
        data = [["JOURS"] + time_slots]
        days_list = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]
        DAYS_MAPPING = {"LUNDI": 1, "MARDI": 2, "MERCREDI": 3, "JEUDI": 4, "VENDREDI": 5, "SAMEDI": 6}
        SLOT_TO_HOUR = {"09h00-10h30": 9, "10h45-12h15": 10, "12h30-14h00": 12, "14h15-15h45": 14, "16h00-17h30": 16}

        for day_name in days_list:
            row = [day_name]
            day_idx = DAYS_MAPPING.get(day_name)
            
            for slot in time_slots:
                if day_name == "VENDREDI" and slot == "14h15-15h45":
                    start_h = 15
                else:
                    start_h = SLOT_TO_HOUR.get(slot, int(slot.split('h')[0]))
                
                cursor.execute("""
                    SELECT s.name as subject, r.name as room, g.name as group_name
                    FROM timetable t
                    JOIN subjects s ON t.course_id = s.id
                    JOIN rooms r ON t.room_id = r.id
                    JOIN groups g ON t.group_id = g.id
                    WHERE t.instructor_id = ? AND t.day = ? AND t.start_hour = ?
                """, (self.instructor_id, day_idx, start_h))
                
                results = cursor.fetchall()
                if results:
                    cell_text = "\n".join([f"{r['subject']}\n{r['group_name']}\n({r['room']})" for r in results])
                    row.append(cell_text)
                else:
                    row.append("")
            data.append(row)
        
        conn.close()

        # Table Style
        table = Table(data, colWidths=[80] + [135] * len(time_slots))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (0, -1), colors.whitesmoke),
        ]))
        elements.append(table)
        
        doc.build(elements)
        return f"PDF exporté vers: {unique_filename}"

    def export_my_timetable_excel(self, filename="Mon_Planning_Enseignant.xlsx"):
        """Export teacher timetable to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            import os
        except ImportError:
            return "Erreur: openpyxl non installé"
            
        # Setup paths
        base_dir = os.getcwd()
        exports_dir = os.path.join(base_dir, "exports")
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
        unique_filename = os.path.join(exports_dir, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mon Planning"
        
        # Styles
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_header = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Headers
        time_slots = ["09h00-10h30", "10h45-12h15", "12h30-14h00", "14h15-15h45", "16h00-17h30"]
        ws.append(["JOURS"] + time_slots)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill_header
            cell.alignment = center_align
            cell.border = thin_border
            
        # Data
        days_list = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]
        DAYS_MAPPING = {"LUNDI": 1, "MARDI": 2, "MERCREDI": 3, "JEUDI": 4, "VENDREDI": 5, "SAMEDI": 6}
        SLOT_TO_HOUR = {"09h00-10h30": 9, "10h45-12h15": 10, "12h30-14h00": 12, "14h15-15h45": 14, "16h00-17h30": 16}
        
        conn = getConnection()
        cursor = conn.cursor()
        
        for day_name in days_list:
            row_data = [day_name]
            day_idx = DAYS_MAPPING.get(day_name)
            
            for slot in time_slots:
                if day_name == "VENDREDI" and slot == "14h15-15h45":
                    start_h = 15
                else:
                    start_h = SLOT_TO_HOUR.get(slot, int(slot.split('h')[0]))
                
                cursor.execute("""
                    SELECT s.name as subject, r.name as room, g.name as group_name
                    FROM timetable t
                    JOIN subjects s ON t.course_id = s.id
                    JOIN rooms r ON t.room_id = r.id
                    JOIN groups g ON t.group_id = g.id
                    WHERE t.instructor_id = ? AND t.day = ? AND t.start_hour = ?
                """, (self.instructor_id, day_idx, start_h))
                
                results = cursor.fetchall()
                if results:
                    text = "\n".join([f"{r['subject']}\n{r['group_name']}\n({r['room']})" for r in results])
                    row_data.append(text)
                else:
                    row_data.append("")
            
            ws.append(row_data)
        
        conn.close()
        
        # Formatting rows
        for row in ws.iter_rows(min_row=2, max_row=len(days_list)+1):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border
        
        # Dimensions
        ws.column_dimensions['A'].width = 15
        for col in range(2, 7):
             ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
             
        wb.save(unique_filename)
        return f"Excel exporté vers: {unique_filename}"

    def export_my_timetable_image(self, filename="Mon_Planning_Enseignant.png"):
        """Export teacher timetable to Image (PNG)"""
        try:
            from PIL import Image, ImageDraw, ImageFont
            import os
        except ImportError:
            return "Erreur: Pillow non installé"
            
        # Setup paths
        base_dir = os.getcwd()
        exports_dir = os.path.join(base_dir, "exports")
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
        unique_filename = os.path.join(exports_dir, filename)
        
        # Config
        time_slots = ["09h00-10h30", "10h45-12h15", "12h30-14h00", "14h15-15h45", "16h00-17h30"]
        days_list = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]
        DAYS_MAPPING = {"LUNDI": 1, "MARDI": 2, "MERCREDI": 3, "JEUDI": 4, "VENDREDI": 5, "SAMEDI": 6}
        SLOT_TO_HOUR = {"09h00-10h30": 9, "10h45-12h15": 10, "12h30-14h00": 12, "14h15-15h45": 14, "16h00-17h30": 16}
        
        cell_width = 160
        cell_height = 90
        header_height = 80
        col_header_height = 40
        day_col_width = 100
        
        img_width = day_col_width + (len(time_slots) * cell_width) + 20
        img_height = header_height + col_header_height + (len(days_list) * cell_height) + 20
        
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Fonts
        try:
            font_title = ImageFont.truetype("arial.ttf", 20)
            font_header = ImageFont.truetype("arial.ttf", 12)
            font_cell = ImageFont.truetype("arial.ttf", 10)
        except:
            font_title = ImageFont.load_default()
            font_header = ImageFont.load_default()
            font_cell = ImageFont.load_default()
            
        # Title
        draw.text((20, 20), "Mon Emploi du Temps - FST Tanger", fill='black', font=font_title)
        
        start_x = 10
        start_y = header_height
        
        # Header Row
        draw.rectangle([start_x, start_y, start_x + day_col_width, start_y + col_header_height], fill='#4472C4', outline='black')
        draw.text((start_x + 10, start_y + 10), "JOURS", fill='white', font=font_header)
        
        for i, slot in enumerate(time_slots):
            x = start_x + day_col_width + (i * cell_width)
            draw.rectangle([x, start_y, x + cell_width, start_y + col_header_height], fill='#4472C4', outline='black')
            draw.text((x + 10, start_y + 10), slot, fill='white', font=font_header)
            
        # Data Rows
        conn = getConnection()
        cursor = conn.cursor()
        
        for row_idx, day_name in enumerate(days_list):
            y = start_y + col_header_height + (row_idx * cell_height)
            
            # Day Column
            draw.rectangle([start_x, y, start_x + day_col_width, y + cell_height], fill='#D9E2F3', outline='black')
            draw.text((start_x + 10, y + 35), day_name, fill='black', font=font_header)
            
            day_idx = DAYS_MAPPING.get(day_name)
            
            for col_idx, slot in enumerate(time_slots):
                x = start_x + day_col_width + (col_idx * cell_width)
                
                if day_name == "VENDREDI" and slot == "14h15-15h45":
                    start_h = 15
                else:
                    start_h = SLOT_TO_HOUR.get(slot, int(slot.split('h')[0]))
                    
                cursor.execute("""
                    SELECT s.name as subject, r.name as room, g.name as group_name
                    FROM timetable t
                    JOIN subjects s ON t.course_id = s.id
                    JOIN rooms r ON t.room_id = r.id
                    JOIN groups g ON t.group_id = g.id
                    WHERE t.instructor_id = ? AND t.day = ? AND t.start_hour = ?
                """, (self.instructor_id, day_idx, start_h))
                
                results = cursor.fetchall()
                draw.rectangle([x, y, x + cell_width, y + cell_height], outline='black')
                
                if results:
                    text_lines = []
                    for r in results:
                        text_lines.append(f"{r['subject']}")
                        text_lines.append(f"{r['group_name']} - {r['room']}")
                    
                    text_y = y + 10
                    for line in text_lines:
                        draw.text((x + 5, text_y), line, fill='black', font=font_cell)
                        text_y += 15
        
        conn.close()
        img.save(unique_filename)
        return f"Image exportée vers: {unique_filename}"